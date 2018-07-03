from cravat.cravat_report import CravatReport
import sys
import os
import datetime
import openpyxl
import openpyxl.worksheet.table
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font

class Reporter(CravatReport):
    def setup (self):    
        self.wb = openpyxl.Workbook()
        self.rowno = 1
        self.colno = 1
        self.sheetno = 0
        self.wb.remove(self.wb['Sheet'])
        self.rightborder = Border(
            right=Side(border_style='thin', color='555555'))
        self.bottomborder = Border(
            bottom=Side(border_style='thin', color='555555'))
        self.grayfill = PatternFill(start_color='e0e0e0', 
                                    end_color='e0e0e0', 
                                    fill_type='solid')
        self.boldfont = Font(size=14, bold=True)
        
        self.write_info_sheet()
        
    def end (self):
        if self.savepath == None:
            self.savepath = 'cravat_result.xlsx'
        else: 
            if self.savepath[-5:] != '.xlsx':
                self.savepath = self.savepath + '.xlsx'
        self.wb.save(self.savepath)
    
    def write_info_sheet (self):
        self.rowno = 1
        self.ws = self.wb.create_sheet(index=self.sheetno, title='info')
        self.sheetno += 1
        lines = ['CRAVAT Report', 
            'Created at ' + 
                datetime.datetime.now().strftime('%A %m/%d/%Y %X'),
            ]
        self.colno = 1
        for i in range(len(lines)):
            self.ws[self.get_cell_coordinate()] = lines[i]
            self.rowno += 1
            
    def get_cell_coordinate (self):
        return get_column_letter(self.colno) + str(self.rowno)
    
    def get_cell (self):
        return self.ws.cell(row=self.rowno, column=self.colno)
    
    def write_preface (self, level):
        self.rowno = 1
        self.ws = self.wb.create_sheet(index=self.sheetno, title=level)
        self.sheetno += 1
        
        cell = self.ws.cell(row=3, column=2)
        self.ws.freeze_panes = cell
    
    def write_header (self, level):
        self.colno = 1
        self.lastcols = []
        groupno = 0
        self.graycols = []
        for colgroup in self.colinfo[level]['colgroups']:
            count = colgroup['count']
            if count == 0:
                continue
            groupno += 1
            displayname = colgroup['displayname']
            lastcol = colgroup['lastcol']
            self.ws[self.get_cell_coordinate()] = displayname
            self.ws.merge_cells(
                start_row=self.rowno, 
                start_column=self.colno,
                end_row=self.rowno,
                end_column=self.colno + count - 1)
            cell = self.get_cell()
            cell.font = self.boldfont
            if groupno % 2 == 0:
                for i in range(self.colno, self.colno + count):
                    self.graycols.append(i)
                    cell.fill = self.grayfill
            self.colno += count
            self.lastcols.append(lastcol)
            cell = self.ws.cell(row=self.rowno, column=self.colno-1)
            cell.border = self.rightborder
        self.rowno += 1
        self.colno = 1
        for column in self.colinfo[level]['columns']:
            self.ws[self.get_cell_coordinate()] = column['col_title']
            cell = self.get_cell()
            cell.border = self.bottomborder
            if self.colno in self.lastcols:
                cell.border += self.rightborder
            if self.colno in self.graycols:
                cell.fill = self.grayfill
            self.colno += 1
        self.rowno += 1
        
    def write_table_row (self, row):
        row = [v if v != None else '.' for v in list(row)]
        self.colno = 1
        for cellvalue in row:
            if type(cellvalue) == type('') and cellvalue[:5] == 'http:':
                cellvalue = '=HYPERLINK("' + cellvalue + '", "' +\
                'Link' + '")'
            self.ws[self.get_cell_coordinate()] = cellvalue
            cell = self.get_cell()
            if self.colno in self.lastcols:
                cell.border = self.rightborder
            if self.colno in self.graycols:
                cell.fill = self.grayfill
            self.colno += 1
        self.rowno += 1

def main ():
    r = Reporter(sys.argv)
    r.run()
    
def test ():
    reporter = Reporter([
        '', 'd:\\git\\cravat-newarch\\tmp\\job\\in1000.sqlite',
        '-s', 'd:\\git\\cravat-newarch\\tmp\\job\\in1000.xlsx'])
    reporter.run()
    reporter = Reporter([
        '', 'd:\\git\\cravat-newarch\\tmp\\job\\in1000.sqlite',
        '--filterstring', '{"variant": {"thousandgenomes__af": ">0.1"}}',
        '-s', 'd:\\git\\cravat-newarch\\tmp\\job\\in1000.filtered.xlsx'])
    reporter.run()

if __name__ == '__main__':
    main()
    #test()
